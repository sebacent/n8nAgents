#!/usr/bin/env python3
"""Genera resúmenes anuales en Excel a partir de un CSV de gastos personales.

El script lee un CSV con las columnas ``fecha,descripcion,categoria,monto``,
limpia y valida los datos y produce **un archivo Excel por año** presente en
los datos (o solo el solicitado con ``--año``). Cada archivo contiene:

  * una hoja ``Resumen Anual`` con el pivote mes x categoría, los totales
    mensuales y tres gráficos (distribución por categoría, comparativo entre
    meses y barras apiladas categoría/mes);
  * una hoja por cada mes con movimientos, nombrada en español (Enero,
    Febrero, …), con el detalle de transacciones, el resumen por categoría
    del mes, el total del mes y dos gráficos (pastel y top de categorías).

Uso:
    python resumen_gastos.py --entrada gastos.csv [--salida-dir out] [--año 2026]
"""

from __future__ import annotations

import argparse
import os
import sys
from io import BytesIO

import matplotlib

# Backend no interactivo: el script debe funcionar en servidores sin pantalla.
matplotlib.use("Agg")

import matplotlib.pyplot as plt  # noqa: E402  (debe ir tras matplotlib.use)
import numpy as np  # noqa: E402
import pandas as pd  # noqa: E402
from openpyxl.drawing.image import Image as ImagenExcel  # noqa: E402

# Columnas que el CSV debe contener obligatoriamente (en minúsculas).
COLUMNAS_REQUERIDAS = ["fecha", "descripcion", "categoria", "monto"]

# Columnas opcionales con sus valores por defecto, para compatibilidad hacia
# atrás con CSVs viejos que no las traen.
COLUMNAS_OPCIONALES = {"tipo": "gasto", "fuente": "manual", "id_doc": ""}

# Tipos válidos para la columna 'tipo'.
TIPOS_VALIDOS = {"gasto", "ingreso"}

# Nombres de mes en español, usados como nombres de hoja.
NOMBRES_MESES = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}

# Abreviaturas de mes para ejes de gráficos.
MESES_CORTOS = {
    1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr",
    5: "May", 6: "Jun", 7: "Jul", 8: "Ago",
    9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic",
}


class ErrorDatosGastos(Exception):
    """Error de negocio con un mensaje claro para el usuario final."""


# ---------------------------------------------------------------------------
# Carga y validación
# ---------------------------------------------------------------------------
def cargar_csv(ruta: str) -> pd.DataFrame:
    """Lee el CSV de ``ruta`` y devuelve un DataFrame.

    Normaliza los nombres de columna (minúsculas y sin espacios) y traduce los
    errores técnicos a mensajes comprensibles.
    """
    if not os.path.isfile(ruta):
        raise ErrorDatosGastos(f"No se encontró el archivo CSV: '{ruta}'.")

    try:
        df = pd.read_csv(ruta)
    except pd.errors.EmptyDataError:
        raise ErrorDatosGastos(f"El archivo '{ruta}' está vacío.")
    except pd.errors.ParserError as exc:
        raise ErrorDatosGastos(f"No se pudo interpretar el CSV '{ruta}': {exc}")
    except UnicodeDecodeError as exc:
        raise ErrorDatosGastos(
            f"Problema de codificación al leer '{ruta}'. "
            f"Guarda el archivo en UTF-8. Detalle: {exc}"
        )

    # Homogeneizar los nombres de columna para el resto del flujo.
    df.columns = [str(c).strip().lower() for c in df.columns]

    if df.empty:
        raise ErrorDatosGastos(f"El archivo '{ruta}' no contiene filas de datos.")

    return df


def validar_columnas(df: pd.DataFrame) -> None:
    """Verifica que estén todas las columnas requeridas; si no, aborta."""
    faltantes = [c for c in COLUMNAS_REQUERIDAS if c not in df.columns]
    if faltantes:
        raise ErrorDatosGastos(
            f"Faltan columnas requeridas en el CSV: {', '.join(faltantes)}. "
            f"Se esperaban: {', '.join(COLUMNAS_REQUERIDAS)}."
        )


# ---------------------------------------------------------------------------
# Limpieza de datos
# ---------------------------------------------------------------------------
def limpiar_datos(df: pd.DataFrame) -> pd.DataFrame:
    """Convierte tipos, descarta filas inválidas y añade la columna ``mes``.

    Las fechas no parseables y los montos no numéricos se convierten a valores
    nulos; toda fila con fecha inválida, monto inválido o categoría vacía se
    reporta (con su número de línea en el CSV) y se descarta. Si tras la
    limpieza no queda ninguna fila válida, se aborta.
    """
    df = df.copy()

    # Normalizar texto antes de validar.
    for col in ("descripcion", "categoria"):
        df[col] = df[col].astype("string").str.strip()

    # Conversión tolerante a errores: lo no convertible queda como nulo.
    df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
    df["monto"] = pd.to_numeric(df["monto"], errors="coerce")

    # Detectar el motivo de invalidez de cada fila.
    motivos: dict[int, list[str]] = {}
    for idx in df.index:
        razones: list[str] = []
        if pd.isna(df.at[idx, "fecha"]):
            razones.append("fecha inválida")
        if pd.isna(df.at[idx, "monto"]):
            razones.append("monto no numérico")
        categoria = df.at[idx, "categoria"]
        if pd.isna(categoria) or categoria == "":
            razones.append("categoría vacía")
        if razones:
            motivos[idx] = razones

    # Avisar de cada fila descartada. El +2 mapea el índice 0-based a la línea
    # del CSV (1 por el encabezado + 1 por el inicio en 0).
    for idx in sorted(motivos):
        print(
            f"  Advertencia: fila {idx + 2} descartada "
            f"({', '.join(motivos[idx])}).",
            file=sys.stderr,
        )

    df_valido = df.drop(index=list(motivos)).copy()
    if df_valido.empty:
        raise ErrorDatosGastos(
            "No quedaron filas válidas tras la limpieza; revisa los datos de entrada."
        )

    # Descripción vacía es aceptable: se completa para evitar nulos en la hoja.
    df_valido["descripcion"] = df_valido["descripcion"].fillna("")
    # Etiqueta de periodo "YYYY-MM" usada para agrupar.
    df_valido["mes"] = df_valido["fecha"].dt.strftime("%Y-%m")

    # Completar columnas opcionales (tipo/fuente/id_doc) con sus defaults para
    # mantener compatibilidad con CSVs viejos.
    for col, default in COLUMNAS_OPCIONALES.items():
        if col not in df_valido.columns:
            df_valido[col] = default
        else:
            df_valido[col] = df_valido[col].fillna(default)

    # Normalizar y validar 'tipo'.
    df_valido["tipo"] = df_valido["tipo"].astype(str).str.strip().str.lower()
    invalidos = ~df_valido["tipo"].isin(TIPOS_VALIDOS)
    if invalidos.any():
        valores = sorted(df_valido.loc[invalidos, "tipo"].unique())
        raise ErrorDatosGastos(
            f"Valores inválidos en columna 'tipo': {', '.join(valores)}. "
            f"Se esperaba uno de: {', '.join(sorted(TIPOS_VALIDOS))}."
        )

    columnas = ["fecha", "descripcion", "categoria", "monto", "mes", "tipo", "fuente", "id_doc"]
    return df_valido.sort_values("fecha")[columnas].reset_index(drop=True)


# ---------------------------------------------------------------------------
# Agregaciones
# ---------------------------------------------------------------------------
def particionar_por_año(df: pd.DataFrame) -> dict[int, pd.DataFrame]:
    """Agrupa el DataFrame por año calendario de la fecha."""
    return {
        int(año): grupo.reset_index(drop=True)
        for año, grupo in df.groupby(df["fecha"].dt.year)
    }


def agrupar_por_mes_categoria(df: pd.DataFrame) -> pd.DataFrame:
    """Tabla dinámica mes x categoría (suma) con totales por fila y columna."""
    pivot = pd.pivot_table(
        df,
        index="mes",
        columns="categoria",
        values="monto",
        aggfunc="sum",
        fill_value=0,
    )

    # Añadir totales sin contar dos veces la columna "Total".
    resultado = pivot.copy()
    resultado["Total"] = pivot.sum(axis=1)
    fila_total = pivot.sum(axis=0)
    fila_total["Total"] = float(pivot.to_numpy().sum())
    resultado.loc["Total"] = fila_total
    return resultado


def resumen_mensual(df: pd.DataFrame) -> pd.DataFrame:
    """Total gastado y cantidad de movimientos por mes."""
    resumen = (
        df.groupby("mes")["monto"].agg(total="sum", cantidad="count").reset_index()
    )
    return resumen


def resumen_por_categoria(df_mes: pd.DataFrame) -> pd.DataFrame:
    """Total y porcentaje del mes para cada categoría, ordenado descendente."""
    por_cat = df_mes.groupby("categoria")["monto"].sum().sort_values(ascending=False)
    total_mes = float(por_cat.sum())
    porcentaje = (por_cat / total_mes * 100) if total_mes else por_cat * 0.0
    return pd.DataFrame(
        {
            "categoria": por_cat.index,
            "total": por_cat.to_numpy(),
            "porcentaje": porcentaje.to_numpy().round(2),
        }
    )


# ---------------------------------------------------------------------------
# Gráficos (matplotlib -> PNG en memoria)
# ---------------------------------------------------------------------------
def _etiqueta_mes(mes_str: str) -> str:
    """Convierte '2026-03' en 'Mar'."""
    return MESES_CORTOS[int(mes_str.split("-")[1])]


def _figura_a_imagen(fig) -> BytesIO:
    """Renderiza una figura de matplotlib a un PNG en memoria y la cierra."""
    buffer = BytesIO()
    fig.savefig(buffer, format="png", dpi=120, bbox_inches="tight")
    plt.close(fig)
    buffer.seek(0)
    return buffer


def grafico_distribucion_categorias(df: pd.DataFrame) -> BytesIO:
    """Gráfico de pastel con el peso de cada categoría sobre el gasto total."""
    por_categoria = df.groupby("categoria")["monto"].sum().sort_values(ascending=False)
    total = float(por_categoria.sum())
    pcts = por_categoria / total * 100

    labels_pie = [cat if pct >= 3 else "" for cat, pct in zip(por_categoria.index, pcts)]

    def autopct_fn(pct):
        return f"{pct:.1f}%" if pct >= 3 else ""

    fig, ax = plt.subplots(figsize=(8, 5))
    wedges, _, _ = ax.pie(
        por_categoria.to_numpy(),
        labels=labels_pie,
        autopct=autopct_fn,
        startangle=90,
    )
    ax.axis("equal")
    ax.set_title("Distribución de gastos por categoría")
    ax.legend(wedges, list(por_categoria.index), title="Categoría", loc="center left", bbox_to_anchor=(1, 0.5))
    return _figura_a_imagen(fig)


def grafico_comparativo_meses(df: pd.DataFrame) -> BytesIO:
    """Gráfico de barras con el gasto total de cada mes."""
    por_mes = df.groupby("mes")["monto"].sum()
    x_pos = list(range(len(por_mes)))
    etiquetas = [_etiqueta_mes(m) for m in por_mes.index]

    fig, ax = plt.subplots(figsize=(8, 5))
    ax.bar(x_pos, por_mes.to_numpy(), color="#4C72B0")
    ax.set_title("Comparativo de gastos por mes")
    ax.set_xlabel("Mes")
    ax.set_ylabel("Monto total")
    ax.set_xticks(x_pos)
    ax.set_xticklabels(etiquetas, rotation=0)
    for i, valor in enumerate(por_mes.to_numpy()):
        ax.text(i, valor, f"{valor:,.0f}", ha="center", va="bottom", fontsize=8)
    return _figura_a_imagen(fig)


def grafico_apilado_categoria_mes(df: pd.DataFrame) -> BytesIO:
    """Barras apiladas: aporte de cada categoría dentro de cada mes."""
    pivot = pd.pivot_table(
        df,
        index="mes",
        columns="categoria",
        values="monto",
        aggfunc="sum",
        fill_value=0,
    )

    x_pos = list(range(len(pivot)))
    etiquetas = [_etiqueta_mes(m) for m in pivot.index]

    fig, ax = plt.subplots(figsize=(9, 5))
    acumulado = np.zeros(len(pivot))
    for categoria in pivot.columns:
        valores = pivot[categoria].to_numpy()
        ax.bar(x_pos, valores, bottom=acumulado, label=str(categoria))
        acumulado += valores

    ax.set_title("Gastos por categoría y mes")
    ax.set_xlabel("Mes")
    ax.set_ylabel("Monto total")
    ax.set_xticks(x_pos)
    ax.set_xticklabels(etiquetas, rotation=0)
    ax.legend(title="Categoría", bbox_to_anchor=(1.02, 1), loc="upper left")
    return _figura_a_imagen(fig)


def grafico_pastel_mes(df_mes: pd.DataFrame, nombre_mes: str) -> BytesIO:
    """Pastel de categorías para un único mes."""
    por_cat = df_mes.groupby("categoria")["monto"].sum().sort_values(ascending=False)
    total = float(por_cat.sum())
    pcts = por_cat / total * 100

    labels_pie = [cat if pct >= 3 else "" for cat, pct in zip(por_cat.index, pcts)]

    def autopct_fn(pct):
        return f"{pct:.1f}%" if pct >= 3 else ""

    fig, ax = plt.subplots(figsize=(7, 4.5))
    wedges, _, _ = ax.pie(
        por_cat.to_numpy(),
        labels=labels_pie,
        autopct=autopct_fn,
        startangle=90,
    )
    ax.axis("equal")
    ax.set_title(f"Distribución por categoría · {nombre_mes}")
    ax.legend(wedges, list(por_cat.index), title="Categoría", loc="center left", bbox_to_anchor=(1, 0.5))
    return _figura_a_imagen(fig)


def grafico_top_categorias_mes(df_mes: pd.DataFrame, nombre_mes: str) -> BytesIO:
    """Barras horizontales con las categorías ordenadas por monto."""
    # Orden ascendente: matplotlib pinta el primer elemento abajo, por lo que
    # la categoría con mayor gasto queda en la parte superior del gráfico.
    por_cat = df_mes.groupby("categoria")["monto"].sum().sort_values(ascending=True)

    fig, ax = plt.subplots(figsize=(7, 4.5))
    ax.barh(list(por_cat.index), por_cat.to_numpy(), color="#4C72B0")
    ax.set_title(f"Top categorías · {nombre_mes}")
    ax.set_xlabel("Monto total")
    for i, valor in enumerate(por_cat.to_numpy()):
        ax.text(valor, i, f" {valor:,.0f}", va="center", fontsize=8)
    fig.tight_layout()
    return _figura_a_imagen(fig)


# ---------------------------------------------------------------------------
# Escritura del Excel
# ---------------------------------------------------------------------------
def _autoajustar_columnas(hoja, dataframe: pd.DataFrame, offset: int = 0) -> None:
    """Ajusta el ancho de columna al contenido más largo (con un tope)."""
    for posicion, columna in enumerate(dataframe.columns, start=1 + offset):
        valores = dataframe[columna].astype(str)
        ancho = max([len(str(columna)), *(len(v) for v in valores)]) + 2
        hoja.column_dimensions[hoja.cell(row=1, column=posicion).column_letter].width = (
            min(ancho, 40)
        )


def escribir_hoja_mes(writer, df_mes: pd.DataFrame, nombre_hoja: str) -> None:
    """Detalle + resumen por categoría + total + dos gráficos de un mes."""
    # 1) Detalle de transacciones (sin la columna interna 'mes').
    detalle = df_mes[["fecha", "descripcion", "categoria", "monto"]].copy()
    detalle["fecha"] = detalle["fecha"].dt.date
    detalle.to_excel(writer, sheet_name=nombre_hoja, index=False, startrow=0)
    hoja = writer.sheets[nombre_hoja]
    _autoajustar_columnas(hoja, detalle)

    # 2) Resumen por categoría con una fila en blanco de separación.
    resumen_cat = resumen_por_categoria(df_mes)
    inicio_resumen = len(detalle) + 3
    resumen_cat.to_excel(
        writer, sheet_name=nombre_hoja, index=False, startrow=inicio_resumen
    )

    # 3) Total del mes, justo debajo del resumen por categoría.
    fila_total_excel = inicio_resumen + len(resumen_cat) + 2  # 1-indexed
    hoja.cell(row=fila_total_excel, column=1, value="Total del mes")
    hoja.cell(row=fila_total_excel, column=2, value=float(df_mes["monto"].sum()))

    # 4) Gráficos del mes a la derecha de las tablas.
    hoja.add_image(ImagenExcel(grafico_pastel_mes(df_mes, nombre_hoja)), "F1")
    hoja.add_image(ImagenExcel(grafico_top_categorias_mes(df_mes, nombre_hoja)), "F26")


def escribir_hoja_ingresos(writer, df_ingresos: pd.DataFrame) -> None:
    """Hoja con el detalle de ingresos del año + total por mes (sin gráficos)."""
    detalle = df_ingresos[["fecha", "descripcion", "categoria", "monto", "fuente"]].copy()
    detalle["fecha"] = detalle["fecha"].dt.date
    detalle.to_excel(writer, sheet_name="Ingresos", index=False)
    hoja = writer.sheets["Ingresos"]
    _autoajustar_columnas(hoja, detalle)

    # Total por mes debajo.
    por_mes = (
        df_ingresos.groupby("mes")["monto"]
        .agg(total="sum", cantidad="count")
        .reset_index()
    )
    inicio = len(detalle) + 3
    por_mes.to_excel(writer, sheet_name="Ingresos", index=False, startrow=inicio)


def escribir_excel_anual(df_año: pd.DataFrame, año: int, ruta_salida: str) -> None:
    """Genera el Excel de un año: hoja Resumen Anual + una hoja por mes.

    Los gráficos y las agregaciones usan solo movimientos con ``tipo=gasto``.
    Si hay ingresos en el año, se agrega una hoja ``Ingresos`` con su detalle.
    """
    df_gastos = df_año[df_año["tipo"] == "gasto"].reset_index(drop=True)
    df_ingresos = df_año[df_año["tipo"] == "ingreso"].reset_index(drop=True)

    if df_gastos.empty:
        raise ErrorDatosGastos(
            f"El año {año} no tiene movimientos con tipo='gasto'; no se puede "
            "generar el resumen anual."
        )

    pivot = agrupar_por_mes_categoria(df_gastos)
    resumen_mes = resumen_mensual(df_gastos)

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        # Hoja 1: Resumen Anual (pivote + tabla mensual + 3 gráficos).
        pivot.to_excel(writer, sheet_name="Resumen Anual")
        hoja_resumen = writer.sheets["Resumen Anual"]
        inicio_resumen = len(pivot) + 3
        resumen_mes.to_excel(
            writer,
            sheet_name="Resumen Anual",
            startrow=inicio_resumen,
            index=False,
        )
        hoja_resumen.add_image(
            ImagenExcel(grafico_distribucion_categorias(df_gastos)), "J1"
        )
        hoja_resumen.add_image(
            ImagenExcel(grafico_comparativo_meses(df_gastos)), "J28"
        )
        hoja_resumen.add_image(
            ImagenExcel(grafico_apilado_categoria_mes(df_gastos)), "J55"
        )

        # Hojas mensuales en orden cronológico, solo meses con gastos.
        meses_con_datos = sorted(int(m) for m in df_gastos["fecha"].dt.month.unique())
        for num_mes in meses_con_datos:
            df_mes = df_gastos[df_gastos["fecha"].dt.month == num_mes].reset_index(drop=True)
            escribir_hoja_mes(writer, df_mes, NOMBRES_MESES[num_mes])

        # Hoja opcional de ingresos al final, solo si hay datos.
        if not df_ingresos.empty:
            escribir_hoja_ingresos(writer, df_ingresos)


# ---------------------------------------------------------------------------
# Punto de entrada
# ---------------------------------------------------------------------------
def parsear_argumentos() -> argparse.Namespace:
    """Define y parsea los argumentos de línea de comandos."""
    parser = argparse.ArgumentParser(
        description=(
            "Genera un Excel por cada año presente en el CSV de gastos personales."
        )
    )
    parser.add_argument(
        "--entrada",
        default="gastos.csv",
        help="Ruta del CSV de entrada (por defecto: gastos.csv).",
    )
    parser.add_argument(
        "--salida-dir",
        dest="salida_dir",
        default=".",
        help="Directorio donde se guardarán los Excel generados (por defecto: el actual).",
    )
    parser.add_argument(
        "--año",
        dest="año",
        type=int,
        default=None,
        help="Procesar solo el año indicado (YYYY). Si no se pasa, se genera uno por año.",
    )
    return parser.parse_args()


def main() -> None:
    """Orquesta el flujo completo: cargar, validar, procesar y exportar."""
    args = parsear_argumentos()
    archivos_generados: list[tuple[str, int]] = []
    try:
        df_crudo = cargar_csv(args.entrada)
        validar_columnas(df_crudo)
        df = limpiar_datos(df_crudo)

        por_año = particionar_por_año(df)

        # Filtrar al año pedido si corresponde, validando existencia.
        if args.año is not None:
            if args.año not in por_año:
                disponibles = ", ".join(str(a) for a in sorted(por_año))
                raise ErrorDatosGastos(
                    f"No hay datos para el año {args.año}. "
                    f"Años disponibles en el CSV: {disponibles}."
                )
            por_año = {args.año: por_año[args.año]}

        os.makedirs(args.salida_dir, exist_ok=True)
        for año in sorted(por_año):
            ruta = os.path.join(args.salida_dir, f"resumen_gastos_{año}.xlsx")
            escribir_excel_anual(por_año[año], año, ruta)
            archivos_generados.append((ruta, len(por_año[año])))
    except ErrorDatosGastos as exc:
        print(f"Error: {exc}", file=sys.stderr)
        sys.exit(1)

    print("Listo. Archivos generados:")
    for ruta, n in archivos_generados:
        print(f"  - {ruta} ({n} registros)")


if __name__ == "__main__":
    main()
